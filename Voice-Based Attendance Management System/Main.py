import speech_recognition as sr
from openpyxl import Workbook, load_workbook


def mark_attendance(student_id):
    # Load attendance.xlsx workbook
    try:
        workbook = load_workbook('attendance.xlsx')
    except FileNotFoundError:
        workbook = Workbook()

    # Select the active sheet
    sheet = workbook.active

    # Check if the student ID exists in the first column
    for row in sheet.iter_rows(values_only=True):
        if row[0] == student_id:
            print("Attendance already marked for this student ID.")
            return

    # Add a new row for the student ID and mark attendance as "Present"
    sheet.append([student_id, "Present"])

    # Save the workbook
    workbook.save('attendance.xlsx')
    print("Attendance marked successfully.")


# Initialize the speech recognizer
recognizer = sr.Recognizer()

# Prompt the user to enter the student ID
student_id = input("Enter the student ID: ")

# Use speech recognition to listen for "Present sir"
with sr.Microphone() as source:
    print("Please say 'Present sir' to mark attendance.")
    audio = recognizer.listen(source)

try:
    # Convert speech to text
    speech_text = recognizer.recognize_google(audio)
    if speech_text.lower() == "present sir":
        # Mark attendance
        mark_attendance(student_id)
    else:
        print("Invalid command. Attendance not marked.")
except sr.UnknownValueError:
    print("Speech recognition could not understand audio.")
except sr.RequestError as e:
    print(f"Could not request results from speech recognition service: {e}")
